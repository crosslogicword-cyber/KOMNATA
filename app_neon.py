import glob
from openpyxl import load_workbook
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime, timedelta, timezone
APP_TZ = timezone(timedelta(hours=2))


def now_local_str():
    return datetime.now(APP_TZ).strftime('%Y-%m-%d %H:%M:%S')

def now_local_display():
    return datetime.now(APP_TZ).strftime('%d.%m.%Y %H:%M')

import os
import re
import sqlite3
try:
    import psycopg2
    from psycopg2.extras import RealDictCursor
except Exception:
    psycopg2 = None
    RealDictCursor = None
from contextlib import closing
from difflib import SequenceMatcher
from werkzeug.utils import secure_filename
from flask import Flask, Response, flash, jsonify, redirect, render_template, request, url_for, send_file

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, 'komnata.db')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'komnata-dev-secret')

DEFAULT_SECTORS = [f"{row} {col}" for row in range(1, 7) for col in ["A", "B", "C"]]
DEFAULT_LAYOUTS = [
    {
        'name': 'Standard 0-5 A,B',
        'start_number': 0,
        'end_number': 5,
        'letters': 'A,B',
        'description': 'Domyślny układ startowy dla każdego regału.',
    },
    {
        'name': 'Rozszerzenie 0-5 A,B,C',
        'start_number': 0,
        'end_number': 5,
        'letters': 'A,B,C',
        'description': 'Układ rozszerzony o dodatkowy poziom C.',
    },
]
DEFAULT_RACK_NAMES = [f"Regał {number}" for number in range(1, 7)]
OPTIONAL_SECTOR_NAME = 'Bez sektora'
OPTIONAL_RACK_NAME = 'Bez regału'
NUMERIC_PREFIX_RE = re.compile(r'^(\d+)')



class DBCompatConnection:
    def __init__(self, raw, engine: str):
        self.raw = raw
        self.engine = engine

    def execute(self, query, params=()):
        params = params or ()
        if self.engine == "postgres":
            pg_query = query.replace("%", "%%")
            pg_query = pg_query.replace("?", "%s")
            cur = self.raw.cursor(cursor_factory=RealDictCursor)
            cur.execute(pg_query, params)
            return cur
        return self.raw.execute(query, params)

    def commit(self):
        return self.raw.commit()

    def rollback(self):
        return self.raw.rollback()

    def close(self):
        return self.raw.close()

    def executescript(self, script):
        if self.engine == "postgres":
            cur = self.raw.cursor(cursor_factory=RealDictCursor)
            for part in [x.strip() for x in script.split(';') if x.strip()]:
                cur.execute(part)
            return cur
        return self.raw.executescript(script)

def get_db():
    database_url = (os.environ.get("DATABASE_URL") or "").strip()

    if database_url and psycopg2 is not None:
        raw = psycopg2.connect(database_url, sslmode="require")
        raw.autocommit = False
        return DBCompatConnection(raw, "postgres")

    raw = sqlite3.connect(DB_PATH)
    raw.row_factory = sqlite3.Row
    return DBCompatConnection(raw, "sqlite")


def natural_sort_key(value: str):
    value = value or ''
    match = NUMERIC_PREFIX_RE.match(value)
    if match:
        return (0, int(match.group(1)), value[len(match.group(1)):])
    return (1, value)


def normalize_letters(raw_letters: str) -> str:
    letters = []
    seen = set()
    for chunk in (raw_letters or '').replace(';', ',').split(','):
        letter = chunk.strip().upper()
        if not letter:
            continue
        if not re.fullmatch(r'[A-Z0-9]+', letter):
            continue
        if letter in seen:
            continue
        seen.add(letter)
        letters.append(letter)
    return ','.join(letters)


def build_slot_codes(start_number: int, end_number: int, letters_csv: str):
    letters = [letter.strip().upper() for letter in (letters_csv or '').split(',') if letter.strip()]
    if end_number < start_number:
        start_number, end_number = end_number, start_number
    return [f"{number}{letter}" for number in range(start_number, end_number + 1) for letter in letters]


def ensure_column(conn: sqlite3.Connection, table_name: str, column_name: str, definition: str):
    columns = {row['name'] for row in conn.execute(f"PRAGMA table_info({table_name})").fetchall()}
    if column_name not in columns:
        conn.execute(f"ALTER TABLE {table_name} ADD COLUMN {column_name} {definition}")


def get_layout_by_name(conn: sqlite3.Connection, name: str):
    return conn.execute("SELECT * FROM layouts WHERE name = ?", (name,)).fetchone()


def apply_layout_to_rack(conn: sqlite3.Connection, rack_id: int, layout_row: sqlite3.Row):
    codes = build_slot_codes(layout_row['start_number'], layout_row['end_number'], layout_row['letters'])
    for sort_order, code in enumerate(codes, start=1):
        conn.execute(
            """
            INSERT INTO rack_slots(rack_id, code, is_active, sort_order)
            VALUES (?, ?, 1, ?)
            ON CONFLICT(rack_id, code) DO UPDATE SET
                is_active = 1,
                sort_order = excluded.sort_order
            """,
            (rack_id, code, sort_order),
        )


def fetch_layouts(active_only: bool = True):
    query = "SELECT * FROM layouts"
    if active_only:
        query += " WHERE is_active = TRUE"
    query += " ORDER BY LOWER(name), name"
    with closing(get_db()) as conn:
        return conn.execute(query).fetchall()


def fetch_sectors(active_only: bool = True):
    query = "SELECT * FROM sectors"
    conditions = ["name <> ?"]
    params = [OPTIONAL_SECTOR_NAME]
    if active_only:
        conditions.append("is_active = TRUE")
    query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY LOWER(name), name"
    with closing(get_db()) as conn:
        return conn.execute(query, params).fetchall()


def fetch_racks(active_only: bool = True, only_main_racks: bool = False):
    query = """
        SELECT racks.*, layouts.name AS layout_name
        FROM racks
        LEFT JOIN layouts ON layouts.id = racks.layout_id
    """
    conditions = []
    params = []
    if active_only:
        conditions.append("racks.is_active = TRUE")
    if only_main_racks:
        conditions.append("racks.name LIKE 'Regał %'")
        conditions.append("racks.name <> ?")
        params.append(OPTIONAL_RACK_NAME)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY LOWER(racks.name), racks.name"
    with closing(get_db()) as conn:
        return conn.execute(query, params).fetchall()


def fetch_slots_by_rack(active_only: bool = True):
    query = "SELECT rack_slots.*, racks.name AS rack_name FROM rack_slots JOIN racks ON racks.id = rack_slots.rack_id"
    if active_only:
        query += " WHERE rack_slots.is_active = TRUE"
    query += " ORDER BY LOWER(racks.name), racks.name, rack_slots.sort_order, LOWER(rack_slots.code), rack_slots.code"
    with closing(get_db()) as conn:
        rows = conn.execute(query).fetchall()

    grouped = {}
    for row in rows:
        grouped.setdefault(str(row['rack_id']), []).append({'id': row['id'], 'code': row['code']})
    return grouped


def fetch_slot_counts():
    with closing(get_db()) as conn:
        rows = conn.execute(
            "SELECT rack_id, COUNT(*) AS slot_count FROM rack_slots WHERE is_active = TRUE GROUP BY rack_id"
        ).fetchall()
    return {row['rack_id']: row['slot_count'] for row in rows}


def all_active_slot_codes():
    with closing(get_db()) as conn:
        rows = conn.execute("SELECT DISTINCT code FROM rack_slots WHERE is_active = TRUE").fetchall()
    return sorted((row['code'] for row in rows), key=natural_sort_key)


def item_base_select():
    return """
        SELECT items.id,
               items.product_name,
               items.quantity,
               items.notes,
               items.extra_field,
               items.created_at,
               COALESCE(items.updated_at, items.created_at) AS updated_at,
               to_char(COALESCE(items.updated_at, items.created_at) + interval '2 hours', 'DD.MM.YYYY HH24:MI') AS updated_at_display,
               items.sector_id,
               items.rack_id,
               items.slot_id,
               CASE WHEN sectors.name = 'Bez sektora' THEN '-' ELSE sectors.name END AS sector_name,
               CASE WHEN racks.name = 'Bez regału' THEN '-' ELSE racks.name END AS rack_name,
               rack_slots.code AS slot_code
        FROM items
        JOIN sectors ON sectors.id = items.sector_id
        JOIN racks ON racks.id = items.rack_id
        LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
    """


def record_item_history(conn: sqlite3.Connection, item_id: int | None, action: str, details: str = ''):
    conn.execute(
        "INSERT INTO item_history(item_id, action, details) VALUES (?, ?, ?)",
        (item_id, action, details.strip() or None),
    )


def get_latest_system_activity():
    with closing(get_db()) as conn:
        row = conn.execute(
            "SELECT to_char(changed_at + interval '2 hours', 'DD.MM.YYYY HH24:MI') AS changed_at_display FROM item_history ORDER BY changed_at DESC, id DESC LIMIT 1"
        ).fetchone()
    return row['changed_at_display'] if row and row['changed_at_display'] else None


@app.context_processor
def inject_system_meta():
    return {'latest_system_activity': get_latest_system_activity()}


def resolve_location_ids(conn: sqlite3.Connection, sector_id_raw: str, rack_id_raw: str, slot_id_raw: str):
    sector_id = (sector_id_raw or '').strip()
    rack_id = (rack_id_raw or '').strip()
    slot_id = (slot_id_raw or '').strip()

    if not sector_id and not rack_id:
        return None, None, None, 'Wybierz przynajmniej sektor albo regał z miejscem.'

    if not sector_id:
        fallback_sector = conn.execute(
            "SELECT id FROM sectors WHERE name = ?",
            (OPTIONAL_SECTOR_NAME,),
        ).fetchone()
        sector_id = str(fallback_sector['id']) if fallback_sector else ''
    else:
        sector_exists = conn.execute(
            "SELECT id FROM sectors WHERE id = ? AND is_active = TRUE",
            (sector_id,),
        ).fetchone()
        if sector_exists is None:
            return None, None, None, 'Wybrany sektor nie istnieje albo jest nieaktywny.'

    if not rack_id:
        fallback_rack = conn.execute(
            "SELECT id FROM racks WHERE name = ?",
            (OPTIONAL_RACK_NAME,),
        ).fetchone()
        rack_id = str(fallback_rack['id']) if fallback_rack else ''
        slot_id = ''
    else:
        rack_exists = conn.execute(
            "SELECT id FROM racks WHERE id = ? AND is_active = TRUE",
            (rack_id,),
        ).fetchone()
        if rack_exists is None:
            return None, None, None, 'Wybrany regał nie istnieje albo jest nieaktywny.'
        if not slot_id:
            return None, None, None, 'Jeśli wybierasz regał, wskaż też miejsce w regale.'
        slot_row = conn.execute(
            "SELECT id FROM rack_slots WHERE id = ? AND rack_id = ? AND is_active = TRUE",
            (slot_id, rack_id),
        ).fetchone()
        if slot_row is None:
            return None, None, None, 'Wybrane miejsce nie należy do tego regału albo jest nieaktywne.'

    return int(sector_id), int(rack_id), int(slot_id) if slot_id else None, None


def init_db():
    with closing(get_db()) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS sectors (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS layouts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                start_number INTEGER NOT NULL,
                end_number INTEGER NOT NULL,
                letters TEXT NOT NULL,
                description TEXT,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS racks (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT UNIQUE NOT NULL,
                layout_id INTEGER,
                is_active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(layout_id) REFERENCES layouts(id)
            );

            CREATE TABLE IF NOT EXISTS rack_slots (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rack_id INTEGER NOT NULL,
                code TEXT NOT NULL,
                is_active INTEGER NOT NULL DEFAULT 1,
                sort_order INTEGER NOT NULL DEFAULT 0,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(rack_id, code),
                FOREIGN KEY(rack_id) REFERENCES racks(id)
            );

            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                product_name TEXT NOT NULL,
                quantity TEXT,
                notes TEXT,
                extra_field TEXT,
                sector_id INTEGER NOT NULL,
                rack_id INTEGER NOT NULL,
                slot_id INTEGER,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT,
                FOREIGN KEY(sector_id) REFERENCES sectors(id),
                FOREIGN KEY(rack_id) REFERENCES racks(id),
                FOREIGN KEY(slot_id) REFERENCES rack_slots(id)
            );

            CREATE TABLE IF NOT EXISTS item_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                item_id INTEGER,
                action TEXT NOT NULL,
                details TEXT,
                changed_at TEXT DEFAULT CURRENT_TIMESTAMP
            );
            """
        )

        ensure_column(conn, 'items', 'slot_id', 'INTEGER')
        ensure_column(conn, 'items', 'extra_field', 'TEXT')
        ensure_column(conn, 'items', 'updated_at', 'TEXT')
        ensure_column(conn, 'racks', 'layout_id', 'INTEGER')
        ensure_column(conn, 'rack_slots', 'sort_order', 'INTEGER NOT NULL DEFAULT 0')

        conn.execute("UPDATE items SET updated_at = created_at WHERE updated_at IS NULL")

        for sector_name in DEFAULT_SECTORS:
            conn.execute("INSERT OR IGNORE INTO sectors(name, is_active) VALUES (?, 1)", (sector_name,))
        conn.execute("INSERT OR IGNORE INTO sectors(name, is_active) VALUES (?, 1)", (OPTIONAL_SECTOR_NAME,))

        for layout in DEFAULT_LAYOUTS:
            conn.execute(
                """
                INSERT OR IGNORE INTO layouts(name, start_number, end_number, letters, description, is_active)
                VALUES (?, ?, ?, ?, ?, 1)
                """,
                (
                    layout['name'],
                    layout['start_number'],
                    layout['end_number'],
                    layout['letters'],
                    layout['description'],
                ),
            )

        standard_layout = get_layout_by_name(conn, 'Standard 0-5 A,B')
        conn.execute("INSERT OR IGNORE INTO racks(name, layout_id, is_active) VALUES (?, NULL, 1)", (OPTIONAL_RACK_NAME,))

        existing_rack_count = conn.execute(
            "SELECT COUNT(*) AS count FROM racks WHERE name <> ?",
            (OPTIONAL_RACK_NAME,),
        ).fetchone()['count']
        if existing_rack_count == 0 and standard_layout:
            for rack_name in DEFAULT_RACK_NAMES:
                conn.execute(
                    "INSERT INTO racks(name, layout_id, is_active) VALUES (?, ?, 1)",
                    (rack_name, standard_layout['id']),
                )

        rack_rows = conn.execute("SELECT * FROM racks ORDER BY id").fetchall()
        for rack in rack_rows:
            if rack['name'] == OPTIONAL_RACK_NAME:
                continue
            layout_row = None
            if rack['layout_id']:
                layout_row = conn.execute("SELECT * FROM layouts WHERE id = ?", (rack['layout_id'],)).fetchone()
            else:
                layout_row = standard_layout
                if layout_row:
                    conn.execute("UPDATE racks SET layout_id = ? WHERE id = ?", (layout_row['id'], rack['id']))
            if layout_row:
                apply_layout_to_rack(conn, rack['id'], layout_row)

        conn.commit()


def search_suggestions(term: str, limit: int = 8):
    clean = (term or '').strip().lower()
    if not clean:
        return []

    with closing(get_db()) as conn:
        rows = conn.execute(
            item_base_select() + " ORDER BY LOWER(items.product_name), items.product_name"
        ).fetchall()

    scored = []
    seen = set()
    for row in rows:
        lowered = row['product_name'].lower()
        if lowered in seen:
            continue
        seen.add(lowered)

        score = 0
        if clean in lowered:
            score += 100
        score += int(SequenceMatcher(None, clean, lowered).ratio() * 100)
        if score >= 35:
            scored.append(
                {
                    'product_name': row['product_name'],
                    'sector': row['sector_name'],
                    'rack': row['rack_name'],
                    'slot': row['slot_code'] or '-',
                    'score': score,
                }
            )
    scored.sort(key=lambda item: (-item['score'], item['product_name'].lower()))
    return scored[:limit]



@app.context_processor
def inject_recent_items():
    recent_items = []
    try:
        with closing(get_db()) as conn:
            recent_items = conn.execute(
                '''
                SELECT
                    items.id,
                    items.product_name,
                    items.quantity,
                    items.notes,
                    CASE WHEN sectors.name = ? THEN '-' ELSE sectors.name END AS sector_name,
                    CASE WHEN racks.name = ? THEN '-' ELSE racks.name END AS rack_name,
                    COALESCE(rack_slots.code, '') AS slot_code,
                    to_char(COALESCE(items.updated_at, items.created_at) + interval '2 hours', 'DD.MM.YYYY HH24:MI') AS changed_display
                FROM items
                JOIN sectors ON sectors.id = items.sector_id
                JOIN racks ON racks.id = items.rack_id
                LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
                ORDER BY COALESCE(items.updated_at, items.created_at) DESC, items.id DESC
                LIMIT 10
                ''',
                (OPTIONAL_SECTOR_NAME, OPTIONAL_RACK_NAME)
            ).fetchall()
    except Exception:
        recent_items = []
    return {"recent_items": recent_items}



@app.context_processor
def inject_system_status():
    db_label = 'błąd'
    db_ok = False
    excel_ok = False
    last_modification = ''
    last_export = ''

    try:
        with closing(get_db()) as conn:
            conn_type = (getattr(conn, 'engine', '') or type(conn).__module__.lower())

            if 'psycopg' in conn_type or 'postgres' in conn_type:
                db_label = 'Neon'
                db_ok = True
            elif 'sqlite' in conn_type:
                db_label = 'lokalna SQLite'
                db_ok = True
            else:
                db_label = 'nieznana'
                db_ok = True
    except Exception:
        db_label = 'błąd'
        db_ok = False

    try:
        excel_ok = os.path.exists('komnata_auto_export.xlsx') or os.path.exists('komnata_backup.xlsx')
    except Exception:
        excel_ok = False

    try:
        with closing(get_db()) as conn:
            row = conn.execute(
                '''
                SELECT strftime('%d.%m.%Y %H:%M', MAX(COALESCE(updated_at, created_at)), '+2 hours') AS last_mod
                FROM items
                '''
            ).fetchone()
            if row:
                if hasattr(row, 'keys'):
                    last_modification = row['last_mod']
                else:
                    last_modification = row[0]
    except Exception:
        last_modification = ''

    try:
        export_candidates = glob.glob('komnata_export_*.xlsx')
        if export_candidates:
            latest_export = max(export_candidates, key=lambda p: os.path.getmtime(p))
            last_export = datetime.fromtimestamp(os.path.getmtime(latest_export)).strftime('%d.%m.%Y %H:%M')
    except Exception:
        last_export = ''

    return {
        'system_status': {
            'db_label': db_label,
            'db_ok': db_ok,
            'excel_ok': excel_ok,
            'last_modification': last_modification,
            'last_export': last_export,
        }
    }


@app.route('/')
def dashboard():
    print("[DASH] start", flush=True)
    sector_filter = request.args.get('sector_id', '').strip()
    rack_filter = request.args.get('rack_id', '').strip()
    slot_filter = request.args.get('slot_id', '').strip()
    print(f"[DASH] filters sector={sector_filter!r} rack={rack_filter!r} slot={slot_filter!r}", flush=True)

    with closing(get_db()) as conn:
        print("[DASH] db opened", flush=True)
        sectors = fetch_sectors(active_only=True)
        print(f"[DASH] sectors ok: {len(sectors)}", flush=True)
        racks = fetch_racks(active_only=True)
        print(f"[DASH] racks ok: {len(racks)}", flush=True)
        slot_counts = fetch_slot_counts()
        print(f"[DASH] slot_counts ok: {len(slot_counts)}", flush=True)

        last_added_row = conn.execute(
            '''
            SELECT rack_id, slot_id
            FROM items
            ORDER BY COALESCE(updated_at, created_at) DESC, id DESC
            LIMIT 1
            '''
        ).fetchone()

        add_selected_rack_id = str(last_added_row['rack_id']) if last_added_row and last_added_row['rack_id'] is not None else ''
        add_selected_slot_id = str(last_added_row['slot_id']) if last_added_row and last_added_row['slot_id'] is not None else ''

        print("[DASH] before rack_slots_all", flush=True)
        rack_slots_all = conn.execute(
            '''
            SELECT
                rack_slots.id,
                rack_slots.rack_id,
                rack_slots.code,
                rack_slots.sort_order,
                rack_slots.is_active
            FROM rack_slots
            WHERE rack_slots.is_active = TRUE
            ORDER BY rack_slots.rack_id, rack_slots.sort_order, rack_slots.code
            '''
        ).fetchall()
        print(f"[DASH] rack_slots_all ok: {len(rack_slots_all)}", flush=True)

        slots_by_rack = {}
        for row in rack_slots_all:
            rack_key = str(row['rack_id'])
            slots_by_rack.setdefault(rack_key, []).append({
                'id': row['id'],
                'code': row['code']
            })

        query = '''
            SELECT
                items.id,
                items.product_name,
                items.quantity,
                items.notes,
                items.extra_field,
                CASE WHEN sectors.name = ? THEN '-' ELSE sectors.name END AS sector_name,
                CASE WHEN racks.name = ? THEN '-' ELSE racks.name END AS rack_name,
                COALESCE(rack_slots.code, '') AS slot_code,
                to_char(COALESCE(items.updated_at, items.created_at) + interval '2 hours', 'DD.MM.YYYY HH24:MI') AS updated_at_display,
                COALESCE(items.updated_at, items.created_at) AS sort_ts
            FROM items
            JOIN sectors ON sectors.id = items.sector_id
            JOIN racks ON racks.id = items.rack_id
            LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
            WHERE 1=1
        '''
        params = [OPTIONAL_SECTOR_NAME, OPTIONAL_RACK_NAME]

        if sector_filter:
            query += " AND items.sector_id = ?"
            params.append(sector_filter)

        if rack_filter:
            query += " AND items.rack_id = ?"
            params.append(rack_filter)

        if slot_filter:
            query += " AND items.slot_id = ?"
            params.append(slot_filter)

        query += " ORDER BY COALESCE(items.updated_at, items.created_at) DESC, items.id DESC"

        print("[DASH] before items query", flush=True)
        items = conn.execute(query, params).fetchall()
        print(f"[DASH] items ok: {len(items)}", flush=True)

    grouped = {}
    for item in items:
        sector_name = item['sector_name'] or '-'
        rack_name = item['rack_name'] or '-'
        slot_code = item['slot_code'] or ''
        grouped.setdefault((sector_name, rack_name, slot_code), []).append(item)

    print(f"[DASH] grouped ok: {len(grouped)}", flush=True)
    print("[DASH] before render_template", flush=True)
    return render_template(
        'dashboard.html',
        items=items,
        grouped=grouped,
        sectors=sectors,
        racks=racks,
        rack_slots=rack_slots_all,
        slot_counts=slot_counts,
        selected_sector_id=sector_filter,
        selected_rack_id=rack_filter,
        selected_slot_id=slot_filter,
        add_selected_rack_id=add_selected_rack_id,
        add_selected_slot_id=add_selected_slot_id,
        slots_by_rack=slots_by_rack,
    )


@app.route('/items/add', methods=['POST'])
def add_item():
    product_name = request.form.get('product_name', '').strip()
    quantity = request.form.get('quantity', '').strip()
    notes = request.form.get('notes', '').strip()
    extra_field = request.form.get('extra_field', '').strip()
    sector_id_raw = request.form.get('sector_id', '').strip()
    rack_id_raw = request.form.get('rack_id', '').strip()
    slot_id_raw = request.form.get('slot_id', '').strip()

    if not product_name:
        flash('Podaj nazwę produktu.', 'error')
        return redirect(url_for('dashboard'))

    with closing(get_db()) as conn:
        resolved = resolve_location_ids(conn, sector_id_raw, rack_id_raw, slot_id_raw)
        sector_id = resolved[0]
        rack_id = resolved[1]
        slot_id = resolved[2]

        duplicate = conn.execute(
            '''
            SELECT id
            FROM items
            WHERE lower(product_name) = lower(?)
              AND sector_id = ?
              AND rack_id = ?
              AND COALESCE(slot_id, 0) = COALESCE(?, 0)
            LIMIT 1
            ''',
            (product_name, sector_id, rack_id, slot_id)
        ).fetchone()

        if duplicate:
            existing = conn.execute(
                '''
                SELECT quantity
                FROM items
                WHERE id = ?
                ''',
                (duplicate['id'],)
            ).fetchone()

            try:
                old_raw = (existing['quantity'] if existing else None)
                old_qty = int(float(str(old_raw).replace(',', '.'))) if str(old_raw).strip() not in ('', 'None') else 1
            except Exception:
                old_qty = 1

            quantity_to_save = str(old_qty + 1)

            conn.execute(
                '''
                UPDATE items
                SET quantity = ?,
                    notes = CASE
                        WHEN COALESCE(?, '') <> '' THEN ?
                        ELSE notes
                    END,
                    extra_field = CASE
                        WHEN COALESCE(?, '') <> '' THEN ?
                        ELSE extra_field
                    END,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                ''',
                (quantity_to_save, notes, notes, extra_field, extra_field, duplicate['id'])
            )
            conn.commit()

            try:
                save_products_excel_snapshot()
                create_timestamped_excel_backup()
            except Exception:
                pass

            flash('Duplikat wykryty — dopisano 1 sztukę do istniejącego produktu.', 'success')
            return redirect(url_for('dashboard'))

        conn.execute(
            '''
            INSERT INTO items (
                product_name,
                quantity,
                notes,
                extra_field,
                sector_id,
                rack_id,
                slot_id,
                updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''',
            (product_name, quantity, notes, extra_field, sector_id, rack_id, slot_id)
        )
        conn.commit()

    try:
        save_products_excel_snapshot()
        create_timestamped_excel_backup()
    except Exception:
        pass

    flash('Dodano produkt do Komnaty.', 'success')
    return redirect(url_for('dashboard'))


@app.route('/items/<int:item_id>/edit', methods=['GET', 'POST'])
def edit_item(item_id):
    mode = 'edit'
    with closing(get_db()) as conn:
        item = conn.execute(item_base_select() + ' WHERE items.id = ?', (item_id,)).fetchone()
        if item is None:
            flash('Nie znaleziono wskazanego produktu.', 'error')
            return redirect(url_for('search'))

        if request.method == 'POST':
            product_name = request.form.get('product_name', '').strip()
            quantity = request.form.get('quantity', '').strip()
            notes = request.form.get('notes', '').strip()
            extra_field = ''
            sector_id = request.form.get('sector_id', '').strip()
            rack_id = request.form.get('rack_id', '').strip()
            slot_id = request.form.get('slot_id', '').strip()

            if not product_name:
                flash('Uzupełnij nazwę produktu.', 'error')
                return redirect(url_for('edit_item', item_id=item_id, mode=mode))

            sector_id_resolved, rack_id_resolved, slot_id_resolved, error = resolve_location_ids(conn, sector_id, rack_id, slot_id)
            if error:
                flash(error, 'error')
                return redirect(url_for('edit_item', item_id=item_id, mode=mode))

            old_location = f"{item['sector_name']} / {item['rack_name']} / {item['slot_code'] or '-'}"
            new_sector_name = conn.execute("SELECT CASE WHEN name = ? THEN '-' ELSE name END AS name FROM sectors WHERE id = ?", (OPTIONAL_SECTOR_NAME, sector_id_resolved)).fetchone()['name']
            new_rack_name = conn.execute("SELECT CASE WHEN name = ? THEN '-' ELSE name END AS name FROM racks WHERE id = ?", (OPTIONAL_RACK_NAME, rack_id_resolved)).fetchone()['name']
            new_slot_name = '-'
            if slot_id_resolved:
                slot_row = conn.execute("SELECT code FROM rack_slots WHERE id = ?", (slot_id_resolved,)).fetchone()
                new_slot_name = slot_row['code'] if slot_row else '-'
            new_location = f"{new_sector_name} / {new_rack_name} / {new_slot_name}"

            action = 'move' if old_location != new_location else 'edit'
            detail_parts = []
            if action == 'move':
                detail_parts.append(f'Przeniesiono: {old_location} → {new_location}')
            else:
                detail_parts.append(f'Edytowano produkt: {product_name}')

            conn.execute(
                """
                UPDATE items
                SET product_name = ?,
                    quantity = ?,
                    notes = ?,
                    extra_field = ?,
                    sector_id = ?,
                    rack_id = ?,
                    slot_id = ?,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                (product_name, quantity, notes, extra_field, sector_id_resolved, rack_id_resolved, slot_id_resolved, item_id),
            )
            record_item_history(conn, item_id, action, '; '.join(detail_parts))
            conn.commit()

            flash('Pozycja została zaktualizowana.', 'success')
            return redirect(url_for('search', q=product_name))

    return render_template(
        'item_form.html',
        item=item,
        sectors=fetch_sectors(),
        racks=fetch_racks(),
        slots_by_rack=fetch_slots_by_rack(),
        mode=mode,
    )


@app.route('/items/<int:item_id>/delete', methods=['POST'])
def delete_item(item_id):
    redirect_target = request.form.get('next') or request.referrer or url_for('dashboard')
    with closing(get_db()) as conn:
        item = conn.execute(item_base_select() + ' WHERE items.id = ?', (item_id,)).fetchone()
        if item is None:
            flash('Nie znaleziono pozycji do usunięcia.', 'error')
            return redirect(redirect_target)
        record_item_history(conn, item_id, 'delete', f"Usunięto produkt: {item['product_name']} z {item['sector_name']} / {item['rack_name']} / {item['slot_code'] or '-'}")
        conn.execute('DELETE FROM items WHERE id = ?', (item_id,))
        conn.commit()
    flash('Pozycja została usunięta.', 'success')
    return redirect(redirect_target)


@app.route('/search')
def search():
    query = request.args.get('q', '').strip()
    results = []
    suggestions = []
    if query:
        with closing(get_db()) as conn:
            results = conn.execute(
                item_base_select() + " WHERE items.product_name LIKE ? OR items.notes LIKE ? ORDER BY COALESCE(items.updated_at, items.created_at) DESC, LOWER(items.product_name), items.product_name",
                (f'%{query}%', f'%{query}%'),
            ).fetchall()
        suggestions = search_suggestions(query)
    return render_template('search.html', query=query, results=results, suggestions=suggestions)


@app.route('/api/suggestions')
def api_suggestions():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify([])

    like = f'%{q}%'
    with closing(get_db()) as conn:
        rows = conn.execute(
            '''
            SELECT
                items.product_name,
                items.quantity,
                items.notes,
                to_char(COALESCE(items.updated_at, items.created_at) + interval '2 hours', 'DD.MM.YYYY HH24:MI') AS updated_at_display,
                CASE WHEN sectors.name = ? THEN '-' ELSE sectors.name END AS sector,
                CASE WHEN racks.name = ? THEN '-' ELSE racks.name END AS rack,
                COALESCE(rack_slots.code, '') AS slot
            FROM items
            JOIN sectors ON sectors.id = items.sector_id
            JOIN racks ON racks.id = items.rack_id
            LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
            WHERE
                lower(items.product_name) LIKE lower(?)
                OR lower(COALESCE(items.notes, '')) LIKE lower(?)
                OR lower(COALESCE(sectors.name, '')) LIKE lower(?)
                OR lower(COALESCE(racks.name, '')) LIKE lower(?)
                OR lower(COALESCE(rack_slots.code, '')) LIKE lower(?)
            ORDER BY COALESCE(items.updated_at, items.created_at) DESC, LOWER(items.product_name), items.product_name
            LIMIT 20
            ''',
            (
                OPTIONAL_SECTOR_NAME,
                OPTIONAL_RACK_NAME,
                like, like, like, like, like
            )
        ).fetchall()

    return jsonify([
        {
            'product_name': row['product_name'] if hasattr(row, 'keys') else row[0],
            'quantity': row['quantity'] if hasattr(row, 'keys') else row[1],
            'notes': row['notes'] if hasattr(row, 'keys') else row[2],
            'updated_at_display': row['updated_at_display'] if hasattr(row, 'keys') else row[3],
            'sector': row['sector'] if hasattr(row, 'keys') else row[4],
            'rack': row['rack'] if hasattr(row, 'keys') else row[5],
            'slot': row['slot'] if hasattr(row, 'keys') else row[6],
        }
        for row in rows
    ])


@app.route('/storage')
def storage_map():
    selected_sector = request.args.get('sector', '').strip()
    selected_rack = request.args.get('rack', '').strip()
    selected_slot = request.args.get('slot', '').strip()

    query = item_base_select()
    conditions = []
    params = []
    if selected_sector:
        conditions.append('sectors.name = ?')
        params.append(selected_sector)
    if selected_rack:
        conditions.append('racks.name = ?')
        params.append(selected_rack)
    if selected_slot:
        conditions.append('rack_slots.code = ?')
        params.append(selected_slot)
    if conditions:
        query += ' WHERE ' + ' AND '.join(conditions)
    query += ' ORDER BY LOWER(sectors.name), sectors.name, LOWER(racks.name), racks.name, LOWER(rack_slots.code), rack_slots.code, items.created_at ASC, LOWER(items.product_name), items.product_name'

    with closing(get_db()) as conn:
        rows = conn.execute(query, params).fetchall()

    grouped = {}
    for row in rows:
        grouped.setdefault((row['sector_name'], row['rack_name'], row['slot_code'] or '-'), []).append(row)

    return render_template(
        'storage.html',
        grouped=grouped,
        sectors=fetch_sectors(),
        racks=fetch_racks(),
        selected_sector=selected_sector,
        selected_rack=selected_rack,
        selected_slot=selected_slot,
        all_slot_codes=all_active_slot_codes(),
    )


@app.route('/print')
def print_view():
    with closing(get_db()) as conn:
        items = conn.execute(
            '''
            SELECT
                items.id,
                items.product_name,
                items.quantity,
                items.notes,
                CASE WHEN sectors.name = ? THEN '-' ELSE sectors.name END AS sector_name,
                CASE WHEN racks.name = ? THEN '-' ELSE racks.name END AS rack_name,
                COALESCE(rack_slots.code, '') AS slot_code,
                to_char(COALESCE(items.updated_at, items.created_at) + interval '2 hours', 'DD.MM.YYYY HH24:MI') AS added_display,
                COALESCE(items.updated_at, items.created_at) AS sort_ts
            FROM items
            JOIN sectors ON sectors.id = items.sector_id
            JOIN racks ON racks.id = items.rack_id
            LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
            ORDER BY COALESCE(items.updated_at, items.created_at) DESC, items.id DESC
            ''',
            (OPTIONAL_SECTOR_NAME, OPTIONAL_RACK_NAME)
        ).fetchall()

    return render_template('print.html', items=items)


@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        item_type = request.form.get('item_type', '').strip()

        if item_type == 'sector':
            name = request.form.get('name', '').strip()
            if not name:
                flash('Podaj nazwę nowego sektora.', 'error')
                return redirect(url_for('settings'))
            with closing(get_db()) as conn:
                conn.execute('INSERT OR IGNORE INTO sectors(name, is_active) VALUES (?, 1)', (name,))
                conn.commit()
            save_products_excel_snapshot()

            flash('Dodano nowy sektor.', 'success')
            return redirect(url_for('settings'))

        if item_type == 'layout':
            name = request.form.get('layout_name', '').strip()
            description = request.form.get('description', '').strip()
            start_number_raw = request.form.get('start_number', '').strip()
            end_number_raw = request.form.get('end_number', '').strip()
            letters = normalize_letters(request.form.get('letters', ''))
            if not name or not start_number_raw or not end_number_raw or not letters:
                flash('Podaj nazwę układu, zakres numerów i litery miejsc.', 'error')
                return redirect(url_for('settings'))
            try:
                start_number = int(start_number_raw)
                end_number = int(end_number_raw)
            except ValueError:
                flash('Zakres numerów musi składać się z liczb całkowitych.', 'error')
                return redirect(url_for('settings'))

            with closing(get_db()) as conn:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO layouts(name, start_number, end_number, letters, description, is_active)
                    VALUES (?, ?, ?, ?, ?, 1)
                    """,
                    (name, start_number, end_number, letters, description),
                )
                conn.commit()
            save_products_excel_snapshot()

            flash('Dodano nowy własny układ regału.', 'success')
            return redirect(url_for('settings'))

        if item_type == 'rack':
            name = request.form.get('name', '').strip()
            layout_id = request.form.get('layout_id', '').strip()
            if not name:
                flash('Podaj nazwę nowego regału.', 'error')
                return redirect(url_for('settings'))
            with closing(get_db()) as conn:
                layout_row = None
                if layout_id:
                    layout_row = conn.execute('SELECT * FROM layouts WHERE id = ?', (layout_id,)).fetchone()
                if layout_row is None:
                    layout_row = get_layout_by_name(conn, 'Standard 0-5 A,B')
                conn.execute(
                    'INSERT OR IGNORE INTO racks(name, layout_id, is_active) VALUES (?, ?, 1)',
                    (name, layout_row['id'] if layout_row else None),
                )
                rack_row = conn.execute('SELECT * FROM racks WHERE name = ?', (name,)).fetchone()
                if rack_row and layout_row:
                    apply_layout_to_rack(conn, rack_row['id'], layout_row)
                conn.commit()
            save_products_excel_snapshot()

            flash('Dodano nowy regał i przypisano mu układ miejsc.', 'success')
            return redirect(url_for('settings'))

        if item_type == 'apply_layout':
            rack_id = request.form.get('rack_id', '').strip()
            layout_id = request.form.get('layout_id', '').strip()
            if not rack_id or not layout_id:
                flash('Wybierz regał i układ do zastosowania.', 'error')
                return redirect(url_for('settings'))
            with closing(get_db()) as conn:
                rack_row = conn.execute('SELECT * FROM racks WHERE id = ?', (rack_id,)).fetchone()
                layout_row = conn.execute('SELECT * FROM layouts WHERE id = ?', (layout_id,)).fetchone()
                if rack_row is None or layout_row is None:
                    flash('Nie znaleziono wybranego regału lub układu.', 'error')
                    return redirect(url_for('settings'))
                apply_layout_to_rack(conn, int(rack_id), layout_row)
                conn.execute('UPDATE racks SET layout_id = ? WHERE id = ?', (layout_id, rack_id))
                conn.commit()
            save_products_excel_snapshot()

            flash('Układ został zastosowany do regału. Brakujące miejsca zostały dodane.', 'success')
            return redirect(url_for('settings'))

        flash('Nieznana akcja ustawień.', 'error')
        return redirect(url_for('settings'))

    return render_template(
        'settings.html',
        sectors=fetch_sectors(active_only=False),
        racks=fetch_racks(active_only=False),
        layouts=fetch_layouts(active_only=False),
        slot_counts=fetch_slot_counts(),
        slot_codes_default=build_slot_codes(0, 5, 'A,B'),
        slot_codes_c=build_slot_codes(0, 5, 'C'),
    )




def save_products_excel_snapshot():
    wb = Workbook()

    def write_sheet(ws, title, rows):
        ws.title = title
        rows = list(rows)
        if not rows:
            ws.append(['Brak danych'])
            return
        first = rows[0]
        headers = list(first.keys()) if hasattr(first, 'keys') else list(dict(first).keys())
        ws.append(headers)
        for row in rows:
            if hasattr(row, 'keys'):
                row_dict = {k: row[k] for k in row.keys()}
            elif isinstance(row, dict):
                row_dict = row
            else:
                row_dict = dict(row)
            ws.append([row_dict.get(h) for h in headers])

    with closing(get_db()) as conn:
        products = conn.execute(
            '''
            SELECT
                items.id,
                items.product_name,
                items.quantity,
                items.notes,
                CASE WHEN sectors.name = ? THEN '' ELSE sectors.name END AS sector_name,
                CASE WHEN racks.name = ? THEN '' ELSE racks.name END AS rack_name,
                COALESCE(rack_slots.code, '') AS slot_code,
                COALESCE(items.extra_field, '') AS extra_field,
                COALESCE(to_char(items.created_at + interval '2 hours', 'YYYY-MM-DD HH24:MI:SS'), '') AS created_at,
                COALESCE(to_char(items.updated_at + interval '2 hours', 'YYYY-MM-DD HH24:MI:SS'), '') AS updated_at
            FROM items
            JOIN sectors ON sectors.id = items.sector_id
            JOIN racks ON racks.id = items.rack_id
            LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
            ORDER BY items.id DESC
            ''',
            (OPTIONAL_SECTOR_NAME, OPTIONAL_RACK_NAME)
        ).fetchall()

        sectors = conn.execute(
            "SELECT id, name, is_active, created_at FROM sectors ORDER BY name"
        ).fetchall()

        racks = conn.execute(
            '''
            SELECT
                racks.id,
                racks.name,
                racks.is_active,
                racks.created_at,
                racks.layout_id,
                COALESCE(layouts.name, '') AS layout_name
            FROM racks
            LEFT JOIN layouts ON layouts.id = racks.layout_id
            ORDER BY racks.name
            '''
        ).fetchall()

        layouts = conn.execute(
            '''
            SELECT
                id, name, start_number, end_number, letters, description, is_active, created_at
            FROM layouts
            ORDER BY name
            '''
        ).fetchall()

        slots = conn.execute(
            '''
            SELECT
                rack_slots.id,
                rack_slots.rack_id,
                racks.name AS rack_name,
                rack_slots.code,
                rack_slots.sort_order,
                rack_slots.is_active,
                rack_slots.created_at
            FROM rack_slots
            JOIN racks ON racks.id = rack_slots.rack_id
            ORDER BY racks.name, rack_slots.sort_order, rack_slots.code
            '''
        ).fetchall()

        history = conn.execute(
            '''
            SELECT
                item_history.id,
                item_history.item_id,
                item_history.action,
                item_history.details,
                item_history.changed_at,
                COALESCE(items.product_name, '') AS product_name
            FROM item_history
            LEFT JOIN items ON items.id = item_history.item_id
            ORDER BY item_history.id DESC
            '''
        ).fetchall()

    ws = wb.active
    write_sheet(ws, 'Produkty', products)
    write_sheet(wb.create_sheet(), 'Sektory', sectors)
    write_sheet(wb.create_sheet(), 'Regały', racks)
    write_sheet(wb.create_sheet(), 'Układy', layouts)
    write_sheet(wb.create_sheet(), 'Miejsca', slots)
    write_sheet(wb.create_sheet(), 'Historia', history)

    wb.save('komnata_auto_export.xlsx')





def create_timestamped_excel_backup():
    try:
        save_products_excel_snapshot()
        source = 'komnata_auto_export.xlsx'
        target = 'komnata_backup.xlsx'
        with open(source, 'rb') as src:
            data = src.read()
        with open(target, 'wb') as dst:
            dst.write(data)
    except Exception:
        pass

@app.route('/export/backup-now')
def backup_now():
    save_products_excel_snapshot()
    timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    source = 'komnata_auto_export.xlsx'
    target = f'komnata_export_{timestamp}.xlsx'

    with open(source, 'rb') as src:
        data = src.read()
    with open(target, 'wb') as dst:
        dst.write(data)

    return send_file(
        target,
        as_attachment=True,
        download_name=target,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export/auto-excel')
def download_auto_excel():
    save_products_excel_snapshot()
    return send_file(
        'komnata_auto_export.xlsx',
        as_attachment=True,
        download_name='komnata_auto_export.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export/excel')
def export_excel():
    wb = Workbook()

    def write_sheet(ws, title, rows):
        ws.title = title
        rows = list(rows)
        if not rows:
            ws.append(['Brak danych'])
            return

        def row_to_dict(row):
            if hasattr(row, 'keys'):
                return {key: row[key] for key in row.keys()}
            if isinstance(row, dict):
                return row
            return dict(row)

        first = row_to_dict(rows[0])
        headers = list(first.keys())
        ws.append(headers)

        for row in rows:
            row_dict = row_to_dict(row)
            ws.append([row_dict[h] if h in row_dict else None for h in headers])

    with closing(get_db()) as conn:
        products = conn.execute(
            '''
            SELECT
                items.id,
                items.product_name,
                items.quantity,
                items.notes,
                items.extra_field,
                items.created_at,
                items.updated_at,
                CASE WHEN sectors.name = ? THEN '' ELSE sectors.name END AS sector_name,
                CASE WHEN racks.name = ? THEN '' ELSE racks.name END AS rack_name,
                COALESCE(rack_slots.code, '') AS slot_code
            FROM items
            JOIN sectors ON sectors.id = items.sector_id
            JOIN racks ON racks.id = items.rack_id
            LEFT JOIN rack_slots ON rack_slots.id = items.slot_id
            ORDER BY items.id DESC
            ''',
            (OPTIONAL_SECTOR_NAME, OPTIONAL_RACK_NAME)
        ).fetchall()

        sectors = conn.execute(
            "SELECT id, name, is_active, created_at FROM sectors ORDER BY name"
        ).fetchall()

        racks = conn.execute(
            '''
            SELECT
                racks.id,
                racks.name,
                racks.is_active,
                racks.created_at,
                racks.layout_id,
                COALESCE(layouts.name, '') AS layout_name
            FROM racks
            LEFT JOIN layouts ON layouts.id = racks.layout_id
            ORDER BY racks.name
            '''
        ).fetchall()

        layouts = conn.execute(
            '''
            SELECT
                id, name, start_number, end_number, letters, description, is_active, created_at
            FROM layouts
            ORDER BY name
            '''
        ).fetchall()

        slots = conn.execute(
            '''
            SELECT
                rack_slots.id,
                rack_slots.rack_id,
                racks.name AS rack_name,
                rack_slots.code,
                rack_slots.sort_order,
                rack_slots.is_active,
                rack_slots.created_at
            FROM rack_slots
            JOIN racks ON racks.id = rack_slots.rack_id
            ORDER BY racks.name, rack_slots.sort_order, rack_slots.code
            '''
        ).fetchall()

        history = conn.execute(
            '''
            SELECT
                item_history.id,
                item_history.item_id,
                item_history.action,
                item_history.details,
                item_history.changed_at,
                COALESCE(items.product_name, '') AS product_name
            FROM item_history
            LEFT JOIN items ON items.id = item_history.item_id
            ORDER BY item_history.id DESC
            '''
        ).fetchall()

    ws = wb.active
    write_sheet(ws, 'Produkty', products)
    write_sheet(wb.create_sheet(), 'Sektory', sectors)
    write_sheet(wb.create_sheet(), 'Regały', racks)
    write_sheet(wb.create_sheet(), 'Układy', layouts)
    write_sheet(wb.create_sheet(), 'Miejsca', slots)
    write_sheet(wb.create_sheet(), 'Historia', history)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='komnata_export.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )



@app.route('/import/excel', methods=['GET', 'POST'])
def import_excel():
    if request.method == 'POST':
        file = request.files.get('excel_file')
        if not file or not file.filename:
            flash('Wybierz plik Excel.', 'error')
            return redirect(url_for('import_excel'))

        filename = secure_filename(file.filename or 'import.xlsx')
        temp_path = f"_tmp_{filename}"
        file.save(temp_path)

        imported = 0
        skipped = 0

        try:
            wb = load_workbook(temp_path, data_only=True)
            if 'Produkty' not in wb.sheetnames:
                flash('Plik nie zawiera arkusza „Produkty”.', 'error')
                return redirect(url_for('import_excel'))

            ws = wb['Produkty']
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                flash('Arkusz „Produkty” jest pusty.', 'error')
                return redirect(url_for('import_excel'))

            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            index = {name: i for i, name in enumerate(headers)}

            def val(row, name, default=''):
                i = index.get(name)
                if i is None or i >= len(row):
                    return default
                return row[i] if row[i] is not None else default

            with closing(get_db()) as conn:
                for row in rows[1:]:
                    product_name = str(val(row, 'Produkt', '')).strip()
                    quantity = str(val(row, 'Ilość', '')).strip()
                    notes = str(val(row, 'Notatki', '')).strip()
                    sector_name = str(val(row, 'Sektor', '')).strip()
                    rack_name = str(val(row, 'Regał', '')).strip()
                    slot_code = str(val(row, 'Miejsce', '')).strip()

                    if not product_name:
                        skipped += 1
                        continue

                    if not sector_name or sector_name == '-':
                        sector_name = OPTIONAL_SECTOR_NAME
                    if not rack_name or rack_name == '-':
                        rack_name = OPTIONAL_RACK_NAME

                    sector = conn.execute(
                        "SELECT id FROM sectors WHERE name = ? LIMIT 1",
                        (sector_name,)
                    ).fetchone()
                    if not sector:
                        sector = conn.execute(
                            "INSERT INTO sectors(name, is_active) VALUES (?, 1) RETURNING id",
                            (sector_name,)
                        ).fetchone()

                    rack = conn.execute(
                        "SELECT id FROM racks WHERE name = ? LIMIT 1",
                        (rack_name,)
                    ).fetchone()
                    if not rack:
                        rack = conn.execute(
                            "INSERT INTO racks(name, is_active) VALUES (?, 1) RETURNING id",
                            (rack_name,)
                        ).fetchone()

                    slot_id = None
                    if slot_code:
                        slot = conn.execute(
                            "SELECT id FROM rack_slots WHERE rack_id = ? AND code = ? LIMIT 1",
                            (rack['id'], slot_code)
                        ).fetchone()
                        if not slot:
                            slot = conn.execute(
                                "INSERT INTO rack_slots(rack_id, code, is_active, sort_order) VALUES (?, ?, 1, 0) RETURNING id",
                                (rack['id'], slot_code)
                            ).fetchone()
                        slot_id = slot['id']

                    duplicate = conn.execute(
                        '''
                        SELECT id FROM items
                        WHERE lower(product_name) = lower(?)
                          AND sector_id = ?
                          AND rack_id = ?
                          AND COALESCE(slot_id, 0) = COALESCE(?, 0)
                        LIMIT 1
                        ''',
                        (product_name, sector['id'], rack['id'], slot_id)
                    ).fetchone()

                    if duplicate:
                        skipped += 1
                        continue

                    conn.execute(
                        '''
                        INSERT INTO items(product_name, quantity, notes, extra_field, sector_id, rack_id, slot_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''',
                        (product_name, quantity, notes, '', sector['id'], rack['id'], slot_id)
                    )

                    imported += 1

                conn.commit()

            try:
                save_products_excel_snapshot()
            except Exception:
                pass

            flash(f'Import zakończony. Dodano: {imported}, pominięto: {skipped}.', 'success')
            return redirect(url_for('import_excel'))

        finally:
            try:
                Path(temp_path).unlink(missing_ok=True)
            except Exception:
                pass

    return render_template('import_excel.html')


if __name__ == '__main__':
    db = get_db()
    try:
        if getattr(db, 'engine', '') != 'postgres':
            init_db()
    finally:
        db.close()

    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
